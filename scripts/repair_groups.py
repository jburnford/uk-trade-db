#!/usr/bin/env python3
"""Repair article_group by cross-volume voting (fixes the pervasive
group-mis-assignment root cause).

The parser carries a sticky "current group" that gets stuck on the wrong
section when a header is missed, scattering one commodity across bogus groups
(Raisins filed under CARDS PLAYING / COPPER ORE, and — the bug that started
this — wood rows losing 'WOOD AND TIMBER' so the export dropped them). But the
ARTICLE label is reliable and the same article recurs across 34 volumes x 2
engines, so the correct group is recoverable as the plurality vote once labels
are normalized. Cross-checked against the gold hierarchy where available.

Output:
  reference/article_group_authority.csv   article -> canonical group (review)
Prints fragmentation + change + wood-uniformity metrics.
"""
import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

import duckdb
import openpyxl

import validate_gold as V   # reuse sig()

BASE = Path('/home/jic823/uk_trade_db')
GOLD = "/mnt/c/Users/jic823/Dropbox/2026/Londons-Ghost-Acres-British-Imports-1856-1906.xlsx"


def gnorm(g):
    """Normalized group KEY: uppercase, de-hyphenate line breaks, strip OCR
    tails and parentheticals, collapse. 'Tan- ning or Dyeing' == 'TANNING...'"""
    g = (g or '')
    g = re.sub(r'-\s+', '', g)                 # join 'Tan- ning'
    g = re.sub(r'\(.*?\)', ' ', g)             # drop parentheticals
    g = g.upper()
    g = re.sub(r"[^A-Z ]", ' ', g)
    g = re.sub(r'\s+', ' ', g).strip()
    # drop trailing filler that varies by print
    g = re.sub(r'\b(AND|OF|OTHER|THAN|USED|IN|OR|SUBSTANCES|SORTS)\b', ' ', g)
    return re.sub(r'\s+', ' ', g).strip()


def load_gold_groups():
    """article_sig -> Counter(gold group-token-set) from gold hierarchy."""
    ws = openpyxl.load_workbook(GOLD, read_only=True, data_only=True).active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ci = hdr.index('COMMODITY_NAME')
    gg = defaultdict(Counter)
    seen = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        com = str(r[ci]).strip()
        if com in seen:
            continue
        seen.add(com)
        toks = V.sig(com)
        # group tokens = leading words before the specific article; approximate
        # as the full-name token set (used only as an agreement signal)
        gg[toks] = toks
    return gg


def main():
    con = duckdb.connect(str(BASE / 'db' / 'uk_trade.duckdb'), read_only=True)
    rows = con.execute("""
        SELECT article_group, article, count(*) n FROM (
            SELECT article_group, article FROM country_obs WHERE flow='import'
            UNION ALL
            SELECT article_group, article FROM country_obs_inf WHERE flow='import'
        ) GROUP BY article_group, article
    """).fetchall()

    # a commodity's "article" is sometimes actually a country or a bare
    # qualifier the parser mis-promoted (defect: country-as-article). Exclude
    # the country vocabulary and bare qualifiers so the authority is real
    # commodities only.
    countries = {V.cnorm(r[0]) for r in con.execute(
        "SELECT DISTINCT country_raw FROM country_obs WHERE country_raw IS NOT NULL").fetchall()}
    QUALIFIERS = {'raw', 'wet', 'dry', 'fresh', 'dressed', 'undressed', 'ore',
                  'yarn', 'unmanufactured', 'manufactured', 'unenumerated',
                  'other manufactures', 'of other materials', 'other sorts',
                  'rough', 'refined', 'salted', 'pickled', 'ground', 'crude'}

    def is_real_article(art):
        a = (art or '').strip()
        if not a:
            return False
        if V.cnorm(a) in countries:
            return False
        if a.lower().strip(' .:') in QUALIFIERS:
            return False
        return True

    # per article-sig: vote normalized group (weighted by row count); keep a
    # readable representative (most common raw spelling of the winning key)
    votes = defaultdict(Counter)            # asig -> Counter(gkey)
    rep_raw = defaultdict(lambda: defaultdict(Counter))  # asig -> gkey -> Counter(raw)
    art_example = {}
    cur_group = defaultdict(Counter)        # asig -> Counter(current raw group)
    n_country_as_article = 0
    for grp, art, n in rows:
        if not is_real_article(art):
            if V.cnorm((art or '').strip()) in countries:
                n_country_as_article += n
            continue
        asig = V.sig(art)
        if not asig:
            continue
        gk = gnorm(grp)
        if not gk:
            continue
        votes[asig][gk] += n
        rep_raw[asig][gk][(grp or '').strip()] += n
        cur_group[asig][(grp or '').strip()] += n
        art_example.setdefault(asig, (art or '').strip())

    gold_sets = set(load_gold_groups().values())

    authority = {}
    out_rows = []
    frag_before = []
    n_low = 0
    for asig, c in votes.items():
        total = sum(c.values())
        gk, nk = c.most_common(1)[0]
        frac = nk / total
        rep = rep_raw[asig][gk].most_common(1)[0][0]
        frag_before.append(len(c))
        # gold agreement: does any gold commodity's full token-set contain both
        # this article's tokens and the winning group's tokens?
        gktoks = set(gnorm(rep).lower().split())
        gold_agree = any(set(asig) <= set(gs) and gktoks & set(gs)
                         for gs in gold_sets if len(gs) <= 8)
        if frac < 0.5:
            n_low += 1
        authority[asig] = rep
        out_rows.append([art_example[asig], rep, f'{frac:.2f}', total,
                         len(c), 'yes' if gold_agree else '',
                         'REVIEW' if frac < 0.5 else ''])

    # ---- metrics: how many rows would change group
    changed = same = 0
    for asig, cc in cur_group.items():
        canon = authority.get(asig)
        ck = gnorm(canon)
        for raw, n in cc.items():
            if gnorm(raw) == ck:
                same += n
            else:
                changed += n

    # ---- wood uniformity check
    wood_arts = [a for a in authority if {'fir', 'oak', 'teak', 'hewn', 'sawn',
                 'stave', 'mahogany', 'batten', 'deal', 'lath'} & set(a)]
    wood_wt = sum(1 for a in wood_arts if 'WOOD' in gnorm(authority[a])
                  and 'TIMBER' in gnorm(authority[a]))

    out_rows.sort(key=lambda r: (r[6] != 'REVIEW', -int(r[3])))
    with open(BASE / 'reference' / 'article_group_authority.csv', 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['article', 'canonical_group', 'plurality_frac', 'n_rows',
                    'n_groups_before', 'gold_agree', 'flag'])
        w.writerows(out_rows)

    import statistics
    print(f'article commodities (real): {len(authority):,}')
    print(f'country-as-article rows excluded: {n_country_as_article:,} '
          '(a separate parser defect)')
    print(f'fragmentation before: mean {statistics.mean(frag_before):.2f} groups/'
          f'article, max {max(frag_before)}; after repair: 1 by construction')
    print(f'rows keeping group: {same:,}   rows re-grouped: {changed:,} '
          f'({changed/(same+changed):.1%})')
    print(f'low-confidence (<50% plurality): {n_low} articles (flagged REVIEW)')
    print(f'wood articles: {len(wood_arts)}; canonical group = WOOD AND TIMBER '
          f'for {wood_wt}/{len(wood_arts)}')
    print('-> reference/article_group_authority.csv')


if __name__ == '__main__':
    main()

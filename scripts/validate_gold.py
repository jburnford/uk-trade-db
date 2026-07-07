#!/usr/bin/env python3
"""Validate the pipeline's country-detail extraction against the human gold.

Gold = London's Ghost Acres British Imports 1856-1906 (quinquennial, by
country, single-keyed by hand). Overlaps our OCR corpus at 1876/81/86/91/96
(+1871 via tn_1871). This is the primary external truth; we cross-check three
ways here (gold value, national-total agreement, country-vector permutation).

Compares, at the overlap years, gold vs pipeline (country_year_consensus) at
two levels:
  A. COVERAGE   — for every gold commodity-year, did the pipeline extract a
                  matching commodity at all? (recall; the "missing commodities"
                  the timber pass was blind to)
  B. NATIONAL   — matched commodity-year: gold World total vs pipeline national
                  total (printed TOTAL row if parsed, else sum of country rows).
  C. ATTRIBUTION— matched commodity-year: is the pipeline's per-country vector
                  the SAME MULTISET as gold but permuted (label shift), a value
                  error, or a clean match?

Outputs:
  reports/gold_validation_scorecard.md        headline numbers
  reference/commodity_gold_crosswalk.csv       gold<->pipeline name map (review)
  reports/gold_missing_commodities.csv         gold cells with no pipeline match
  reports/gold_national_mismatch.csv           matched but total off >5%
  reports/gold_attribution_defects.csv         permutation / value errors
"""
import csv
import re
from collections import defaultdict
from pathlib import Path

import duckdb
import openpyxl

BASE = Path('/home/jic823/uk_trade_db')
GOLD = "/mnt/c/Users/jic823/Dropbox/2026/Londons-Ghost-Acres-British-Imports-1856-1906.xlsx"
OVERLAP = [1871, 1876, 1881, 1886, 1891, 1896]

# ---------------------------------------------------------------- normalizers
STOP = {'and', 'or', 'of', 'the', 'other', 'sorts', 'sort', 'for', 'to', 'be',
        'used', 'in', 'as', 'not', 'kinds', 'kind', 'a', 'an', 'all', 'than',
        'being', 'made', 'with', 'no', 'otherwise', 'unenumerated', 'unenumd'}


# archaic/variant spellings and compounds -> a common token so gold and
# pipeline names line up (flaxseed=flax seed, pease=peas, indian corn=maize...)
SYN = {'flaxseed': 'flax', 'pease': 'pea', 'peas': 'pea', 'linseed': 'linseed',
       'maize': 'maize', 'cornmeal': 'corn', 'oatmeal': 'oat', 'gutch': 'cutch',
       'myrabolans': 'myrabolan', 'succades': 'succade', 'spermaceiti': 'spermaceti',
       'woons': 'wood', 'saltpetre': 'saltpetre', 'nitre': 'saltpetre'}


def sig(s):
    """Commodity signature: order-insensitive content-token tuple."""
    s = (s or '').lower()
    s = re.sub(r'\(.*?\)', ' ', s)              # drop parentheticals
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    toks = []
    for t in s.split():
        if t in STOP or not t:
            continue
        if len(t) > 4 and t.endswith('s'):       # naive singular
            t = t[:-1]
        t = SYN.get(t, t)
        toks.append(t)
    return tuple(sorted(set(toks)))


def cnorm(s):
    """Country key: both sides come from the same returns vocabulary."""
    s = (s or '').lower().strip()
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    alias = {
        'united states of america': 'united states', 'usa': 'united states',
        'british north america': 'british north america',
        'dominion of canada': 'british north america', 'canada': 'british north america',
        'holland': 'netherlands',
        'russia northern ports': 'russia', 'russia southern ports': 'russia',
    }
    return alias.get(s, s)


def jaccard(a, b):
    sa, sb = set(a), set(b)
    return len(sa & sb) / len(sa | sb) if (sa or sb) else 0.0


def unit_ok(gu, pu):
    """Are the gold and pipeline units the same measure? (loose)."""
    def u(s):
        s = (s or '').lower()
        s = re.sub(r'\(.*?\)', '', s)
        s = re.sub(r'[^a-z]', '', s)
        m = {'cwts': 'cwt', 'cwt': 'cwt', 'tons': 'ton', 'ton': 'ton',
             'loads': 'load', 'load': 'load', 'lbs': 'lb', 'lb': 'lb',
             'qrs': 'qr', 'quarters': 'qr', 'gallons': 'gal', 'gals': 'gal',
             'number': 'no', 'no': 'no', 'pieces': 'pc', 'bushels': 'bu'}
        return m.get(s, s)
    a, b = u(gu), u(pu)
    if not a or not b:
        return True                      # unknown -> don't penalize
    return a == b


# ---------------------------------------------------------------- load gold
def load_gold():
    ws = openpyxl.load_workbook(GOLD, read_only=True, data_only=True).active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    i = {h: k for k, h in enumerate(hdr)}
    Y, C, FROM = i['YEAR'], i['COMMODITY_NAME'], i['FROM_LOCATION']
    UN, AMT, VAL = i['UNIT_NAME'], i['AMOUNT'], i['VALUE']
    over = set(OVERLAP)
    # (commodity, year) -> {country_key: amount}, plus world + unit
    cells = defaultdict(dict)
    world = {}
    unit = {}
    comnames = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        try:
            y = int(r[Y])
        except (TypeError, ValueError):
            continue
        if y not in over:
            continue
        com = str(r[C]).strip()
        loc = str(r[FROM]).strip()
        try:
            amt = float(r[AMT] or 0)
        except (TypeError, ValueError):
            amt = 0.0
        key = (com, y)
        comnames[sig(com)] = com
        unit.setdefault(com, str(r[UN]))
        if loc.lower() == 'world':
            world[key] = amt
        else:
            cells[key][cnorm(loc)] = amt
    return cells, world, unit


# ------------------------------------------------------------ load pipeline
def load_pipeline(con):
    """Group-agnostic: article_group is unreliable (a commodity is scattered
    across bogus groups), so we identify a commodity by its ARTICLE and merge
    its country rows across all groups within a year. Generic articles whose
    sig is empty ('Unenumerated') fall back to the group+article sig."""
    yrs = tuple(OVERLAP)
    rows = con.execute(f"""
        SELECT article_group, article, country, year, quantity, unit
        FROM country_year_consensus
        WHERE flow='import' AND year IN {yrs}
    """).fetchall()
    cells = defaultdict(lambda: defaultdict(float))   # (key_sig, year) -> {country: qty}
    total = defaultdict(float)                         # (key_sig, year) -> printed TOTAL (max)
    unit = defaultdict(lambda: defaultdict(int))       # key_sig -> unit votes
    label = {}                                         # key_sig -> representative name
    for grp, art, ctry, y, q, u in rows:
        y = int(y); q = float(q or 0)
        asig = sig(art)
        ksig = asig if asig else sig(f"{grp or ''} {art or ''}")   # generic -> full
        if not ksig:
            continue
        key = (ksig, y)
        label.setdefault(ksig, (art or '').strip() or f"{grp} {art}")
        if u:
            unit[ksig][u] += 1
        c = cnorm(ctry)
        if c in ('total', 'world'):
            total[key] = max(total[key], q)
        elif ' : ' in (ctry or ''):        # sub-entry: skip (avoid double count)
            continue
        else:
            cells[key][c] += q
    munit = {k: max(v, key=v.get) for k, v in unit.items()}
    return cells, total, munit, label


# ---------------------------------------------------------------- crosswalk
def build_crosswalk(gold_sigs, pipe_sigs):
    """gold_sig -> (pipe_sig, confidence, method).

    Article-subset match: the pipeline key is a commodity's ARTICLE sig, which
    should be a subset of the gold's full (group+article) sig — e.g. article
    (fir,hewn) ⊆ gold (fir,hewn,timber,wood). Score = fraction of the pipeline
    article's tokens covered by the gold name; require the article's rarest
    token present (kills Beef/Beer) and pick the most specific (longest) hit.
    """
    df = defaultdict(int)
    for s in list(gold_sigs) + list(pipe_sigs):
        for t in set(s):
            df[t] += 1
    xwalk = {}
    for gs in gold_sigs:
        if gs in pipe_sigs:
            xwalk[gs] = (gs, 1.0, 'exact')
            continue
        gset = set(gs)
        grare = min(gs, key=lambda t: df.get(t, 0)) if gs else None
        best, bscore, blen = None, 0.0, 0
        for ps in pipe_sigs:
            if not ps:
                continue
            pset = set(ps)
            inter = pset & gset
            # distinctive-token guard: share the rarer of the two sigs' key token
            prare = min(ps, key=lambda t: df.get(t, 0))
            if grare not in pset and prare not in gset:
                continue
            # bidirectional containment: smaller sig mostly inside the larger
            cover = len(inter) / min(len(ps), len(gset))
            if cover >= 0.75 and (cover > bscore or (cover == bscore and len(ps) > blen)):
                best, bscore, blen = ps, cover, len(ps)
        if best is not None:
            xwalk[gs] = (best, round(bscore, 2), 'subset')
        else:
            # last resort: symmetric jaccard on full name
            rare = min(gs, key=lambda t: df.get(t, 0)) if gs else None
            b2, bj = None, 0.0
            for ps in pipe_sigs:
                j = jaccard(gs, ps)
                if j > bj and rare in ps:
                    b2, bj = ps, j
            xwalk[gs] = (b2, round(bj, 2), 'fuzzy') if (b2 and bj >= 0.6) \
                else (None, round(bj, 2), 'unmatched')
    return xwalk


def classify_attribution(gold_vec, pipe_vec):
    """Compare two {country: value} dicts. Returns (label, detail)."""
    def rnd(v):
        return round(v)
    gv = {k: rnd(v) for k, v in gold_vec.items() if rnd(v) > 0}
    pv = {k: rnd(v) for k, v in pipe_vec.items() if rnd(v) > 0}
    if not pv:
        return 'missing', ''
    # per-country agreement (within 1%)
    def close(a, b):
        return abs(a - b) <= max(2, 0.01 * max(a, b))
    matched = sum(1 for k in gv if k in pv and close(gv[k], pv[k]))
    frac = matched / len(gv) if gv else 0
    if frac >= 0.9:
        return 'match', f'{matched}/{len(gv)} countries'
    # permutation test: same value multiset, different country assignment
    from collections import Counter

    def bucket(vals):
        return Counter(round(v, -1) for v in vals)   # tolerance ~10
    gm, pm = bucket(gv.values()), bucket(pv.values())
    inter = sum((gm & pm).values())
    if gv and inter >= 0.8 * len(gv) and frac < 0.6:
        return 'attribution-shift', f'{inter}/{len(gv)} values present but mislabeled'
    return 'value-error', f'{matched}/{len(gv)} countries match'


def main():
    con = duckdb.connect(str(BASE / 'db' / 'uk_trade.duckdb'), read_only=True)
    gcells, gworld, gunit = load_gold()
    pcells, ptotal, punit, plabel = load_pipeline(con)

    gold_sigs = {sig(com) for (com, y) in gcells}
    pipe_sigs = {k for (k, y) in list(pcells) + list(ptotal)}
    xwalk = build_crosswalk(gold_sigs, pipe_sigs)
    gname = {sig(com): com for (com, y) in gcells}

    # ---- iterate gold commodity-years
    cov_matched = cov_missing = 0
    missing_rows, natl_rows, attr_rows, xwalk_rows = [], [], [], []
    natl_exact = natl_close = natl_off = natl_unit = 0
    attr_counts = defaultdict(int)
    per_year = defaultdict(lambda: [0, 0])       # year -> [matched, missing]

    for (com, y), gvec in sorted(gcells.items()):
        gs = sig(com)
        ps, conf, method = xwalk[gs]
        gtot = gworld.get((com, y)) or sum(gvec.values())
        pkey = (ps, y) if ps else None
        pvec = pcells.get(pkey, {}) if pkey else {}
        has = bool(pvec) or (pkey in ptotal)
        if not ps or not has:
            cov_missing += 1
            per_year[y][1] += 1
            note = 'no name match' if not ps else f'name-matched({method}) but absent in {y}'
            missing_rows.append([com, y, gunit.get(com, ''), f'{gtot:.0f}', note])
            continue
        cov_matched += 1
        per_year[y][0] += 1
        ptot = ptotal.get(pkey) or sum(pvec.values())
        if not unit_ok(gunit.get(com, ''), punit.get(ps, '')):
            natl_unit += 1
        elif gtot > 0:
            ratio = ptot / gtot
            if abs(ratio - 1) <= 0.01:
                natl_exact += 1
            elif abs(ratio - 1) <= 0.05:
                natl_close += 1
            else:
                natl_off += 1
                natl_rows.append([com, y, f'{gtot:.0f}', f'{ptot:.0f}',
                                  f'{ratio:.3f}', plabel.get(ps, '')])
        lab, detail = classify_attribution(gvec, pvec)
        attr_counts[lab] += 1
        if lab in ('attribution-shift', 'value-error'):
            attr_rows.append([com, y, lab, detail, plabel.get(ps, '')])

    for gs, (ps, conf, method) in sorted(xwalk.items(), key=lambda kv: gname.get(kv[0], '')):
        xwalk_rows.append([gname.get(gs, ' '.join(gs)),
                           plabel.get(ps, '') if ps else '', method, conf])

    # ---- write outputs
    rep = BASE / 'reports'
    ref = BASE / 'reference'
    _w(rep / 'gold_missing_commodities.csv',
       ['gold_commodity', 'year', 'unit', 'gold_total', 'note'], missing_rows)
    _w(rep / 'gold_national_mismatch.csv',
       ['gold_commodity', 'year', 'gold_total', 'pipe_total', 'ratio', 'pipe_name'],
       sorted(natl_rows, key=lambda r: abs(float(r[4]) - 1), reverse=True))
    _w(rep / 'gold_attribution_defects.csv',
       ['gold_commodity', 'year', 'class', 'detail', 'pipe_name'], attr_rows)
    _w(ref / 'commodity_gold_crosswalk.csv',
       ['gold_commodity', 'pipeline_fullname', 'method', 'confidence'],
       sorted(xwalk_rows))

    tot_cells = cov_matched + cov_missing
    n_gold_com = len({com for (com, y) in gcells})
    matched_com = len({sig(com) for (com, y) in gcells if xwalk[sig(com)][0]})
    lines = [
        '# Gold validation scorecard (Ghost Acres vs pipeline)', '',
        f'Overlap years: {OVERLAP}', '',
        '## A. Commodity coverage',
        f'- Gold commodities (overlap yrs): **{n_gold_com}**; name-matched to a pipeline commodity: **{matched_com}** ({matched_com/n_gold_com:.0%})',
        f'- Gold commodity-year cells: {tot_cells:,} — matched with data **{cov_matched:,}** ({cov_matched/tot_cells:.0%}), missing **{cov_missing:,}** ({cov_missing/tot_cells:.0%})',
        '',
        '## Coverage by year (matched / gold cells)',
    ]
    for y in OVERLAP:
        m, mi = per_year[y]
        tot = m + mi
        lines.append(f'- {y}: matched {m}/{tot} ({m/tot:.0%})' if tot else f'- {y}: —')
    lines += [
        '',
        '## B. National total accuracy (matched cells, same-unit)',
        f'- exact (≤1%): {natl_exact:,}   close (≤5%): {natl_close:,}   off (>5%): {natl_off:,}   unit-differs(skipped): {natl_unit:,}',
        '',
        '## C. Country attribution (matched cells)',
    ]
    for lab in ('match', 'attribution-shift', 'value-error', 'missing'):
        lines.append(f'- {lab}: {attr_counts.get(lab, 0):,}')
    lines += ['', '## Files',
              '- reports/gold_missing_commodities.csv',
              '- reports/gold_national_mismatch.csv',
              '- reports/gold_attribution_defects.csv',
              '- reference/commodity_gold_crosswalk.csv']
    (rep / 'gold_validation_scorecard.md').write_text('\n'.join(lines) + '\n')
    print('\n'.join(lines))


def _w(path, header, rows):
    with open(path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


if __name__ == '__main__':
    main()
